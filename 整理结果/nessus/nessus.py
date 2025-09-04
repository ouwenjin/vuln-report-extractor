import pandas as pd
import os
import datetime
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- 配置 ----------
REFERENCE_FILE = 'Nessus中文报告.xlsx'  # 漏洞引用表
RISK_MAPPING = {'Critical': '紧急', 'High': '高', 'Medium': '中', 'Low': '低', 'None': '无'}

# ---------- CSV处理 ----------
def merge_csv_files():
    csv_files = [f for f in os.listdir(os.getcwd()) if f.endswith('.csv')]
    if not csv_files:
        print("当前目录没有 CSV 文件")
        return pd.DataFrame(), None
    elif len(csv_files) == 1:
        df = pd.read_csv(csv_files[0])
        merged_file = csv_files[0]
    else:
        df_list = [pd.read_csv(f) for f in csv_files]
        df = pd.concat(df_list, ignore_index=True)
        merged_file = 'merged.csv'
        df.to_csv(merged_file, index=False, encoding='utf-8-sig')
        print(f"已合并 {len(csv_files)} 个 CSV 文件为 {merged_file}")
    return df, merged_file

def convert_csv_to_xlsx(csv_file):
    xlsx_file = os.path.splitext(csv_file)[0] + '.xlsx'
    df = pd.read_csv(csv_file)
    df.to_excel(xlsx_file, index=False, sheet_name='ScanData')
    print(f"{csv_file} 已转换为 {xlsx_file}")
    return xlsx_file

# ---------- 引用表 ----------
def load_reference_vuln_table(ref_file, sheet_name='漏洞引用表'):
    try:
        df_ref = pd.read_excel(ref_file, sheet_name=sheet_name, header=0)
        vuln_dict = {}
        for _, row in df_ref.iterrows():
            plugin_id = str(row['编号'])
            vuln_dict[plugin_id] = {
                '英文名称': row.get('漏洞英文名称', ''),
                '中文名称': row.get('漏洞名称', ''),
                '风险等级': row.get('风险等级', ''),
                '漏洞说明': row.get('漏洞说明', ''),
                '加固建议': row.get('加固建议', '')
            }
        return vuln_dict, df_ref
    except Exception as e:
        print(f"引用表加载失败: {e}")
        return {}, pd.DataFrame()

# ---------- 输入数据 ----------
def load_input_data(input_file):
    try:
        xls = pd.ExcelFile(input_file)
        sheet_name = xls.sheet_names[0]
        df_input = pd.read_excel(input_file, sheet_name=sheet_name, header=0)
        df_input.fillna({'CVE':'','Plugin Output':'','Port':'','Synopsis':'',
                         'Description':'','Solution':'','Name':'','Risk':'None'}, inplace=True)
        df_vulns = df_input[df_input['Risk'] != 'None'].copy()
        return df_vulns, df_input['Host'].unique()
    except Exception as e:
        print(f"输入数据加载失败: {e}")
        return pd.DataFrame(), []

# ---------- 扫描结果 ----------
def generate_scan_results(df_vulns, vuln_ref_dict):
    results=[]
    for idx,row in df_vulns.iterrows():
        plugin_id=str(row['Plugin ID'])
        ref=vuln_ref_dict.get(plugin_id,{})
        vuln_name=ref.get('中文名称',row['Name'])
        risk_level=ref.get('风险等级',RISK_MAPPING.get(row['Risk'],'未知'))
        description=ref.get('漏洞说明',row['Synopsis']+'\n'+row['Description'])
        solution=ref.get('加固建议',row['Solution'])
        results.append([row['Host'],row['Port'],vuln_name,risk_level,
                        description,solution,row['CVE'],row['Plugin Output']])
    df_results=pd.DataFrame(results,columns=['IP','端口','漏洞名称','风险等级',
                                             '漏洞说明','加固建议','CVE','扫描返回信息'])
    # 重新生成序号列
    df_results.insert(0, '序号', range(1, len(df_results)+1))
    return df_results

# ---------- 写Excel美化 ----------
def write_scan_results_only(output_file, results_df):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    left_top_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    bold_font = Font(bold=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='扫描结果', index=False)
        ws = writer.sheets['扫描结果']
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = bold_font if cell.row==1 else Font()
                cell.alignment = left_top_align
                cell.border = thin_border
        col_widths_results=[10,20,10,30,10,50,50,20,50]
        for i,w in enumerate(col_widths_results,start=1):
            ws.column_dimensions[get_column_letter(i)].width=w
    print(f"扫描结果生成完成：{output_file}")

# ---------- IP列表 ----------
def export_ip_list(unique_ips, df_vulns):
    ip_file='ip.xlsx'
    df_ips=pd.DataFrame(unique_ips,columns=['IP'])
    df_stats=df_vulns.groupby('Risk').size().reindex(['Critical','High','Medium'],fill_value=0).reset_index()
    df_stats.columns=['风险等级','数量']
    df_stats['风险等级']=df_stats['风险等级'].map(RISK_MAPPING)
    with pd.ExcelWriter(ip_file, engine='openpyxl') as writer:
        df_ips.to_excel(writer, sheet_name='IP列表', index=False)
        df_stats.to_excel(writer, sheet_name='漏洞统计', index=False)
    print(f"IP 列表与漏洞统计输出：{ip_file}")

# ---------- 缺失引用 ----------
def export_missing_reference_examples(df_vulns,vuln_ref_dict,output_file='缺失示例.xlsx'):
    missing_mask=df_vulns['Plugin ID'].astype(str).apply(lambda x: x not in vuln_ref_dict)
    if missing_mask.sum()==0:
        print("没有缺失引用的漏洞。")
        return
    df_missing=df_vulns[missing_mask].copy()
    df_sheet1=df_missing[['Plugin ID','Name','Host','Port','Risk']]
    df_sheet2=df_missing.copy()
    thin_border=Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
    center_align=Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_top_align=Alignment(horizontal='left', vertical='top', wrap_text=True)
    bold_font=Font(bold=True)
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_sheet1.to_excel(writer, sheet_name='缺失引用简要', index=False)
        ws=writer.sheets['缺失引用简要']
        for row in ws.iter_rows(min_row=1,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
            for cell in row:
                cell.font=bold_font if cell.row==1 else Font()
                cell.alignment=center_align
                cell.border=thin_border
        df_sheet2.to_excel(writer, sheet_name='缺失引用样例', index=False)
        ws=writer.sheets['缺失引用样例']
        for row in ws.iter_rows(min_row=1,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
            for cell in row:
                cell.font=bold_font if cell.row==1 else Font()
                cell.alignment=left_top_align
                cell.border=thin_border
    print(f"缺失引用示例已输出：{output_file}, 共 {len(df_missing)} 条记录。")

# ---------- 主流程 ----------
def main():
    df_merged, merged_file=merge_csv_files()
    if df_merged.empty: return
    xlsx_file=convert_csv_to_xlsx(merged_file)
    vuln_ref_dict, ref_df=load_reference_vuln_table(REFERENCE_FILE)
    if df_merged.empty: return
    df_vulns, unique_ips=load_input_data(xlsx_file)
    if df_vulns.empty:
        print("没有漏洞数据，结束。")
        return
    results_df=generate_scan_results(df_vulns,vuln_ref_dict)
    timestamp=datetime.datetime.now().strftime('%Y%m%d%H%M')
    output_file=f"漏洞扫描结果-{timestamp}.xlsx"
    write_scan_results_only(output_file, results_df)
    export_ip_list(unique_ips, df_vulns)
    export_missing_reference_examples(df_vulns, vuln_ref_dict)

    # ---------- 新增：复制到上级目录/整理结果 并改名，只保留中高危 ----------
    high_risk_df = results_df[results_df['风险等级'].isin(['紧急','高''中'])]
    target_dir = os.path.join(os.path.dirname(os.getcwd()), "整理结果")
    os.makedirs(target_dir, exist_ok=True)
    target_file = os.path.join(target_dir, "中高危漏洞.xlsx")
    write_scan_results_only(target_file, high_risk_df)
    print(f"中高危漏洞已输出到：{target_file}")

if __name__=='__main__':
    main()
    
"""
nessus-中文报告转化
迪普-张昊 2025.8.26
"""
