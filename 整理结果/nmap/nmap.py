import os
import re
import shutil
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Font
from tqdm import tqdm
import logging

# ===========================
# 日志配置
# ===========================
LOG_FILE = "merge.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ===========================
# 危险端口和服务定义
# ===========================
dangerous_ports = {
    20,21,23,25,53,69,111,110,2049,143,137,135,139,389,445,161,
    512,513,514,873,1433,1521,1529,3306,3389,5000,5432,
    5900,5901,5902,6379,7001,888,9200,9300,11211,27017,27018
}
dangerous_services = {
    'ftp','telnet','smtp','dns','smb','snmp','rsync','oracle','mysql','mysqlx',
    'mariadb','rdp','postgresql','vnc','redis','weblogic_server','elasticsearch',
    'elasticsearch_transport','memcached','mongodb','mongodb_shard_or_secondary',
    'tftp','nfs','pop3','imap','netbios-ns','msrpc','netbios-ssn','ldap',
    'linux rexec','mssql','oracle db','sybase/db2','ilo','any','oracledb',
    'http','linuxrexec','vnc服务'
}

# ===========================
# 校验 IP
# ===========================
def is_valid_ip(ip):
    if not ip:
        return False
    ipv4_pattern = r"^(25[0-5]|2[0-4]\d|[01]?\d\d?)" \
                   r"(\.(25[0-5]|2[0-4]\d|[01]?\d\d?)){3}$"
    ipv6_pattern = r"^([0-9a-fA-F]{0,4}:){2,7}[0-9a-fA-F]{0,4}$"
    return re.match(ipv4_pattern, ip) or re.match(ipv6_pattern, ip)

# ===========================
# 合并所有 Nmap XML 文件
# ===========================
def merge_all_xml(output_file="out.xml"):
    xml_files = [f for f in os.listdir(".") if f.lower().endswith(".xml")]
    if not xml_files:
        logger.warning("没有找到 XML 文件，跳过合并。")
        return None

    logger.info(f"开始合并 {len(xml_files)} 个 XML 文件 -> {output_file}")
    main_tree = ET.parse(xml_files[0])
    main_root = main_tree.getroot()

    for xml_file in xml_files[1:]:
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            for host in root.findall("host"):
                main_root.append(host)
        except Exception as e:
            logger.error(f"合并文件 {xml_file} 出错: {e}")

    main_tree.write(output_file, encoding="utf-8", xml_declaration=True)
    logger.info(f"XML 合并完成，结果保存为 {output_file}")
    return output_file

# ===========================
# 解析 Nmap XML
# ===========================
def parse_nmap_xml(xml_file):
    results = []
    if not os.path.exists(xml_file):
        logger.warning(f"文件不存在: {xml_file}")
        return results
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        hosts = root.findall("host")
        for h_index, host in enumerate(tqdm(hosts, desc=f"解析Nmap: {xml_file}", unit="host")):
            ip = None
            addr = host.find("address")
            if addr is not None:
                ip = addr.get("addr")

            if not is_valid_ip(ip):
                logger.warning(f"[Nmap] 文件 {xml_file} Host#{h_index} IP 无效: {ip}")

            for port in host.findall(".//port"):
                proto = port.get("protocol")
                portid = port.get("portid")
                state = port.find("state").get("state") if port.find("state") is not None else ""
                service_elem = port.find("service")
                service = service_elem.get("name") if service_elem is not None else ""

                results.append({
                    "IP": ip,
                    "端口/协议": f"{portid}/{proto}",
                    "状态": state,
                    "服务": service,
                    "端口用途": "",
                })
    except Exception as e:
        logger.error(f"解析 Nmap 文件 {xml_file} 出错: {e}")
    return results

# ===========================
# 解析 Excel/CSV 表格
# ===========================
def parse_table(file_path):
    results = []
    if not os.path.exists(file_path):
        logger.error(f"文件不存在: {file_path}")
        return results
    try:
        df = pd.read_excel(file_path) if file_path.lower().endswith(".xlsx") else pd.read_csv(file_path)
        if df.empty:
            logger.warning(f"文件为空: {file_path}")
            return results

        # 列映射
        col_map = {
            "IP": ["IP","ip","地址","Host"],
            "端口/协议": ["端口/协议","端口","Port","port"],
            "状态": ["状态","State","开放状态"],
            "服务": ["服务","Service","协议"],
            "端口用途": ["端口用途","用途","备注","Remark"]
        }
        real_cols = {}
        for std_col, aliases in col_map.items():
            for alias in aliases:
                if alias in df.columns:
                    real_cols[std_col] = alias
                    break
            if std_col not in real_cols:
                real_cols[std_col] = None

        for i, row in tqdm(df.iterrows(), total=len(df), desc=f"解析表格: {file_path}", unit="行"):
            ip = row.get(real_cols["IP"], "") if real_cols["IP"] else ""
            if not is_valid_ip(ip):
                logger.warning(f"[表格] 文件 {file_path} 行 {i+2} IP 无效: {ip}")

            port_proto = row.get(real_cols["端口/协议"], "") if real_cols["端口/协议"] else ""
            if port_proto and " /" not in str(port_proto):
                port_proto = f"{port_proto}/tcp"

            state = row.get(real_cols["状态"], "") if real_cols["状态"] else ""
            service = row.get(real_cols["服务"], "") if real_cols["服务"] else ""
            remark = row.get(real_cols["端口用途"], "") if real_cols["端口用途"] else ""

            results.append({
                "IP": str(ip).strip(),
                "端口/协议": str(port_proto).strip(),
                "状态": str(state).strip(),
                "服务": str(service).strip(),
                "端口用途": str(remark).strip(),
            })
    except Exception as e:
        logger.error(f"解析文件 {file_path} 出错: {e}")
    return results

# ===========================
# 标记危险端口/服务
# ===========================
def mark_dangerous(df):
    def check(row):
        try:
            port = int(str(row["端口/协议"]).split("/")[0])
        except:
            port = None
        service = str(row["服务"]).strip().lower()
        if (port in dangerous_ports) or (service in dangerous_services):
            return "危险端口不允许对外开放"
        return ""
    df["是否必要开放"] = df.apply(check, axis=1)
    return df

# ===========================
# Excel 美化
# ===========================
def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    font = Font(name="宋体", size=12)
    bold_font = Font(name="宋体", size=12, bold=True)
    red_font = Font(name="宋体", size=12, color="FFFF0000")

    column_widths = {"A":36,"B":12,"C":12,"D":18,"E":11,"F":28}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            if cell.row == 1:
                cell.font = bold_font
            if cell.value == "危险端口不允许对外开放":
                cell.font = red_font
    wb.save(file_path)

# ===========================
# 去重逻辑
# ===========================
def auto_dedup(df):
    if df.empty:
        return df, "none"
    before = len(df)
    df.drop_duplicates(subset=["IP","端口/协议","服务","状态","端口用途"], inplace=True)
    after = len(df)
    mode = f"strict ({before-after} 行被删除)"
    return df, mode

# ===========================
# 主函数
# ===========================
def main():
    current_directory = os.getcwd()
    parent_directory = os.path.dirname(current_directory)
    # 明确指定上级目录中的“整理结果”文件夹
    整理结果目录 = os.path.join(parent_directory, "整理结果")

    # 创建上级目录中的“整理结果”文件夹（如果不存在）
    if not os.path.exists(整理结果目录):
        os.makedirs(整理结果目录)
        logger.info(f"创建目录: {整理结果目录}")

    all_results = []

    # 第一步：合并 XML
    merged_xml = merge_all_xml("out.xml")

    # 第二步：解析 Excel/CSV
    input_file = "开放端口.xlsx"
    all_results.extend(parse_table(input_file))

    # 第三步：解析 out.xml
    if merged_xml:
        all_results.extend(parse_nmap_xml(merged_xml))

    if not all_results:
        logger.error("未找到可解析数据。")
        return

    df = pd.DataFrame(all_results)
    df, mode = auto_dedup(df)
    logger.info(f"自动去重模式：{mode}，最终 {len(df)} 行")

    df = mark_dangerous(df)

    output_file = "端口调研表.xlsx"
    df.to_excel(output_file, index=False)
    format_excel(output_file)
    logger.info(f"处理完成，结果保存为 {output_file}")

    # 第四步：移动到上级目录的“整理结果”文件夹
    target_file = os.path.join(整理结果目录, "端口调研表.xlsx")
    try:
        if os.path.exists(target_file):
            os.remove(target_file)
        shutil.move(output_file, target_file)
        logger.info(f"已移动 {output_file} -> {target_file}")
    except Exception as e:
        logger.error(f"移动文件到 {target_file} 失败: {e}")
        return

    # 第五步：清理临时文件
    for tmp_file in ["out.xml", "1.xlsx"]:
        tmp_path = os.path.join(current_directory, tmp_file)
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
                logger.info(f"已删除临时文件: {tmp_path}")
            except Exception as e:
                logger.error(f"删除 {tmp_path} 失败: {e}")

if __name__ == "__main__":
    main()