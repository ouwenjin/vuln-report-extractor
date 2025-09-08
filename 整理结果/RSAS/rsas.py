import os
import sys
import re
import difflib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil
import tempfile

TARGET_COLS = ["序号", "IP", "端口", "漏洞名称", "风险等级", "漏洞说明", "加固建议", "CVE"]

COLUMN_CANDIDATES = {
    "序号": ["序号", "id", "编号", "no", "number"],
    "IP": ["IP", "IP地址", "ipaddress", "ip address", "host"],
    "端口": ["端口", "port"],
    "漏洞名称": ["漏洞名称", "名称", "漏洞", "vuln name", "vulnerability name"],
    "风险等级": ["风险等级", "等级", "risk level", "severity"],
    "漏洞说明": ["漏洞说明", "漏洞描述", "描述", "description"],
    "加固建议": ["加固建议", "整改建议", "修复建议", "建议", "recommendation", "remediation"],
    "CVE": ["CVE", "漏洞CVE编号", "CVE编号", "cve id", "漏洞CVE"]
}

def normalize(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.strip().lower()
    s = re.sub(r'[\s_：:，,。\.\-]+', '', s)
    return s

def find_best_col_for_target(target_keywords, columns):
    norm_map = {normalize(c): c for c in columns}
    norms = list(norm_map.keys())
    for kw in target_keywords:
        nk = normalize(kw)
        if nk in norm_map:
            return norm_map[nk]
    for kw in target_keywords:
        nk = normalize(kw)
        for n in norms:
            if nk and nk in n:
                return norm_map[n]
    for kw in target_keywords:
        best = difflib.get_close_matches(normalize(kw), norms, n=1, cutoff=0.6)
        if best:
            return norm_map[best[0]]
    return None

def is_zhong_or_gao(val):
    if pd.isna(val):
        return False
    s = str(val).strip()
    return bool(re.search(r'[中高]', s))

def ensure_output_dir(script_dir):
    parent = os.path.dirname(script_dir)
    if os.path.basename(script_dir) == "整理结果":
        out = script_dir
    else:
        out = os.path.join(parent, "整理结果")
    os.makedirs(out, exist_ok=True)
    return out

def align_df_to_target(df_src, target_cols, src_columns=None):
    if src_columns is None:
        src_columns = list(df_src.columns)
    aligned = pd.DataFrame()
    for tgt in target_cols:
        candidates = COLUMN_CANDIDATES.get(tgt, [tgt])
        found = find_best_col_for_target(candidates, src_columns)
        if found and found in df_src.columns:
            aligned[tgt] = df_src[found]
        elif tgt in df_src.columns:
            aligned[tgt] = df_src[tgt]
        else:
            aligned[tgt] = ""
    return aligned

def write_and_format_excel(df, path):
    df.to_excel(path, index=False, engine="openpyxl")
    wb = load_workbook(path)
    ws = wb.active

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col = ws.max_column
    max_row = ws.max_row
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    col_name_to_idx = { ws.cell(row=1, column=i).value: i for i in range(1, max_col+1) }
    for cname in ["漏洞说明", "加固建议"]:
        if cname in col_name_to_idx:
            cidx = col_name_to_idx[cname]
            for r in range(2, max_row+1):
                ws.cell(row=r, column=cidx).alignment = wrap
                ws.cell(row=r, column=cidx).border = border

    for r in range(2, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.alignment.wrapText if cell.alignment else False))
            cell.border = border

    ws.freeze_panes = "A2"

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is None:
                l = 0
            else:
                l = len(str(cell.value))
            if l > max_len:
                max_len = l
        adjusted_width = min(max_len + 4, 60)
        ws.column_dimensions[col_letter].width = adjusted_width

    try:
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass

    wb.save(path)

def find_latest_zip(search_dir: str, exclude_name: str = "绿盟.zip"):
    candidates = []
    try:
        for fname in os.listdir(search_dir):
            if not fname.lower().endswith('.zip'):
                continue
            if fname == exclude_name:
                continue
            full = os.path.join(search_dir, fname)
            if os.path.isfile(full):
                candidates.append(full)
    except FileNotFoundError:
        return None
    if not candidates:
        return None
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]

def copy_atomic_to_dest(src_path: str, dest_dir: str, dest_name: str = "绿盟.zip"):
    if not os.path.isfile(src_path):
        raise FileNotFoundError(src_path)
    os.makedirs(dest_dir, exist_ok=True)
    dest_path = os.path.join(dest_dir, dest_name)
    tmp = tempfile.NamedTemporaryFile(delete=False, dir=dest_dir, suffix=".tmp")
    tmp_name = tmp.name
    tmp.close()
    try:
        with open(src_path, 'rb') as fr, open(tmp_name, 'wb') as fw:
            shutil.copyfileobj(fr, fw)
        os.replace(tmp_name, dest_path)
    except Exception:
        try:
            if os.path.exists(tmp_name):
                os.remove(tmp_name)
        except Exception:
            pass
        raise
    return dest_path

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(script_dir, "整理结果", "漏洞报告.xlsx")
    if not os.path.isfile(input_path):
        print(f"[ERROR] 找不到输入文件：{input_path}")
        sys.exit(1)

    output_dir = ensure_output_dir(script_dir)
    sorted_path = os.path.join(output_dir, "漏洞报告_整理.xlsx")
    zhg_path = os.path.join(output_dir, "中高危漏洞.xlsx")

    try:
        df_src = pd.read_excel(input_path, engine="openpyxl")
    except Exception as e:
        print(f"[ERROR] 读取源文件失败：{e}")
        sys.exit(1)

    df_aligned = align_df_to_target(df_src, TARGET_COLS, src_columns=list(df_src.columns))

    df_filtered = df_aligned[df_aligned["风险等级"].apply(is_zhong_or_gao)].copy().reset_index(drop=True)

    def renumber(df):
        if "序号" not in df.columns:
            df.insert(0, "序号", "")
        df["序号"] = range(1, len(df) + 1)
        cols = [c for c in df.columns if c != "序号"]
        df = df[["序号"] + cols]
        return df

    df_filtered = renumber(df_filtered)

    try:
        write_and_format_excel(df_filtered, sorted_path)
        print(f"[*] 已保存整理文件（临时）：{sorted_path}")
    except Exception as e:
        print(f"[ERROR] 保存整理文件失败：{e}")
        sys.exit(1)

    if os.path.isfile(zhg_path):
        try:
            df_exist = pd.read_excel(zhg_path, engine="openpyxl")
            df_exist_aligned = align_df_to_target(df_exist, TARGET_COLS, src_columns=list(df_exist.columns))
            print(f"[*] 读取到了已存在的 中高危文件（{zhg_path}），原行数：{len(df_exist)}，映射后列：{list(df_exist_aligned.columns)}")
        except Exception as e:
            print(f"[WARN] 读取已有 中高危 文件失败，将创建新表：{e}")
            df_exist_aligned = pd.DataFrame(columns=TARGET_COLS)
    else:
        df_exist_aligned = pd.DataFrame(columns=TARGET_COLS)
        print(f"[*] 未找到 中高危漏洞.xlsx，将在 {output_dir} 创建新文件。")

    def drop_all_empty_rows(df):
        if df is None or df.shape[0] == 0:
            return df.copy()
        mask = df.apply(lambda row: any((not pd.isna(v) and (not (isinstance(v, str) and v.strip() == ""))) for v in row), axis=1)
        return df[mask].reset_index(drop=True)

    exist_rows = drop_all_empty_rows(df_exist_aligned)
    new_rows = drop_all_empty_rows(df_filtered)

    try:
        if exist_rows.shape[0] == 0 and new_rows.shape[0] == 0:
            print("[*] 检测到原有中高危文件和本次整理出的文件（除表头）都为空，将写入提示行。")
            row = {c: "" for c in TARGET_COLS}
            row["风险等级"] = "暂未发现中高危漏洞"
            df_out = pd.DataFrame([row], columns=TARGET_COLS)
            df_out = renumber(df_out)
            write_and_format_excel(df_out, zhg_path)
            print(f"[*] 已写入提示行并保存到：{zhg_path}")
        else:
            df_combined = pd.concat([exist_rows, new_rows], ignore_index=True, sort=False)
            df_combined = renumber(df_combined[TARGET_COLS])
            write_and_format_excel(df_combined, zhg_path)
            print(f"[*] 已将追加结果保存到：{zhg_path}（总行数：{len(df_combined)})")
    except Exception as e:
        print(f"[ERROR] 写入 中高危漏洞.xlsx 失败：{e}")
        try:
            if os.path.isfile(sorted_path):
                os.remove(sorted_path)
                print(f"[*] 已删除临时整理文件：{sorted_path}")
        except Exception:
            pass
        sys.exit(1)

    try:
        if os.path.isfile(sorted_path):
            os.remove(sorted_path)
            print(f"[*] 已删除临时整理文件：{sorted_path}")
        else:
            print(f"[*] 临时整理文件未找到（可能已被删除或未生成）：{sorted_path}")
    except Exception as e:
        print(f"[WARN] 删除临时整理文件失败：{e}")

    print("[*] 全部操作完成。")

    try:
        src_zip = find_latest_zip(script_dir, exclude_name="绿盟.zip")
        if not src_zip:
            print(f"[*] 未在脚本目录 ({script_dir}) 发现可用的 zip，跳过生成 绿盟.zip")
            return
        try:
            dest_zip = copy_atomic_to_dest(src_zip, output_dir, dest_name="绿盟.zip")
            print(f"[*] 已将源 zip 复制并重命名为：{dest_zip}")
        except Exception as e:
            print(f"[WARN] 将源 zip 复制为 绿盟.zip 失败：{e}")
    except Exception as e:
        print(f"[WARN] 在尝试生成 绿盟.zip 时发生异常：{e}")

if __name__ == "__main__":
    main()
